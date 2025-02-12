# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image
# import os

# # Caminhos dos arquivos
# input_excel_path = "./PDR410.xlsx"
# # img_path = "./IMAGENS/0131290.BMP"
# output_excel_path = "./teste_planilha.xlsx"

# IMAGE_WIDTH = 80  # Definir a largura da imagemssss
# IMAGE_HEIGHT = 80  # Definir a altura da imagem

# # Carregar o arquivo Excel
# wb = load_workbook(input_excel_path)
# ws = wb.active  # Obtém a primeira planilha

# # Ajustar largura da coluna e altura das linhas para exibir a imagem corretamente
# ws.column_dimensions['A'].width = 15  

# # Percorrer todas as linhas da coluna A (a partir da linha 2)
# for row in range(2, ws.max_row + 1):
#     cell_value = ws[f'A{row}'].value  # Obter o valor da célula
#     print(f"Linha {row}: {cell_value}")  # Log do valor da célula

#     # Se a célula contém um caminho de imagem, substituímos pela imagem desejada
#     if cell_value and cell_value.startswith("./IMAGENS/"):  
#         img_path = cell_value.strip()
#         if os.path.exists(img_path):   
#             img = Image(img_path)  # Criar uma nova instância da imagem
#             img.width = IMAGE_WIDTH
#             img.height = IMAGE_HEIGHT

#             # Definir a posição da imagem no centro da célula (ajuste de deslocamento)
#             cell = ws[f'A{row}']
#             cell_coordinate = cell.coordinate  # Exemplo: "A2"
#             ws.add_image(img, cell_coordinate)  # Inserir imagem na célula

#             # Ajustar altura da linha para exibir corretamente a imagem
#             ws.row_dimensions[row].height = IMAGE_HEIGHT * 0.85

#             # Remover o texto da célula para manter apenas a imagem
#             cell.value = ""
#         else:
#             print(f"Arquivo não existe: {img_path}")  # Log do valor da célula

# # Criar pasta de saída, se não existir
# os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)

# # Salvar o arquivo modificado
# wb.save(output_excel_path)

# print(f"Imagem inserida e arquivo salvo em: {output_excel_path}")

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import sys

# Verifica se o usuário passou o nome do arquivo
if len(sys.argv) < 2:
    print("Uso: python3 excelImage.py <arquivo.xlsx>")
    sys.exit(1)

# Obtém o caminho do arquivo de entrada a partir do argumento do terminal
input_excel_path = sys.argv[1]
output_excel_path = input_excel_path.replace(".xlsx", ".xlsx")

IMAGE_WIDTH = 200  # Definir a largura da imagem
IMAGE_HEIGHT = 200  # Definir a altura da imagem

# Verifica se o arquivo de entrada existe
if not os.path.exists(input_excel_path):
    print(f"Erro: O arquivo '{input_excel_path}' não foi encontrado.")
    sys.exit(1)

# Carregar o arquivo Excel
wb = load_workbook(input_excel_path)
ws = wb.active  # Obtém a primeira planilha

# Ajustar largura da coluna e altura das linhas para exibir a imagem corretamente
ws.column_dimensions['A'].width = 15  

# Percorrer todas as linhas da coluna A (a partir da linha 2)
for row in range(2, ws.max_row + 1):
    cell_value = ws[f'A{row}'].value  # Obter o valor da célula
    print(f"Linha {row}: {cell_value}")  # Log do valor da célula

    # Se a célula contém um caminho de imagem, substituímos pela imagem desejada
    if cell_value and cell_value.startswith("./IMAGENS/"):  
        img_path = cell_value.strip()
        if os.path.exists(img_path):   
            img = Image(img_path)  # Criar uma nova instância da imagem
            img.width = IMAGE_WIDTH
            img.height = IMAGE_HEIGHT

            # Definir a posição da imagem no centro da célula (ajuste de deslocamento)
            cell = ws[f'A{row}']
            cell_coordinate = cell.coordinate  # Exemplo: "A2"
            ws.add_image(img, cell_coordinate)  # Inserir imagem na célula

            # Ajustar altura da linha para exibir corretamente a imagem
            ws.row_dimensions[row].height = IMAGE_HEIGHT * 0.85

            # Remover o texto da célula para manter apenas a imagem
            cell.value = ""
        else:
            print(f"Arquivo não existe: {img_path}")  # Log do valor da célula

# Criar pasta de saída, se necessário (corrigindo o erro)
output_dir = os.path.dirname(output_excel_path)
if output_dir:  # Só cria se existir um diretório especificado
    os.makedirs(output_dir, exist_ok=True)

# Salvar o arquivo modificado
wb.save(output_excel_path)

print(f"Imagem inserida e arquivo salvo em: {output_excel_path}")

