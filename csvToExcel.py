import os, sys
import pandas as pd

from os.path import exists
# from openpyxl import Workbook
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


def ajustarColunas(arquivo_excel):
    wb = load_workbook(arquivo_excel)
    ws = wb.active

    # Ajusta a largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtém a letra da coluna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)  # Ajuste para um espaço extra
        ws.column_dimensions[column].width = adjusted_width

    # Ajusta a altura das linhas
    for row in ws.iter_rows():
        max_height = 15  # Altura mínima padrão
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                lines = cell.value.count("\n") + 1  # Conta quebras de linha
                max_height = max(max_height, lines * 15)  # Ajusta a altura
        ws.row_dimensions[row[0].row].height = max_height  # Aplica a altura ajustada

    wb.save(arquivo_excel)

def adicionarImagensExcel(arquivo_excel):
    # print(f"Adicionando imagens ao arquivo: {arquivo_excel}")

    wb = load_workbook(arquivo_excel)
    ws = wb.active

    IMAGE_WIDTH = 200  # Definir a largura da imagem
    IMAGE_HEIGHT = 200  # Definir a altura da imagem

    SIZE = 128, 128

    # Ajustar largura da coluna e altura das linhas para exibir a imagem corretamente
    ws.column_dimensions['A'].width = 30

    seqAnterior = 0

    # Percorrer todas as linhas da coluna A (a partir da linha 2)
    for row in range(2, ws.max_row + 1):
        seqAtual =ws[f'D{row}'].value

        if seqAtual != seqAnterior:
            cell_value = ws[f'A{row}'].value  # Obter o valor da célul

            # Se a célula contém um caminho de imagem, substituímos pela imagem desejada
            if cell_value:
                img_path = cell_value.strip()
                # img_path = r"D:\Workspace\Python\fcpCSV\IMAGENS\0131290.BMP"
                # print(f"Imagem: {img_path}")
                if os.path.exists(img_path):
                    # print(f"Arquivo existe: {img_path}")  # Log do valor da célula

                    img = Image(img_path)  # Criar uma nova instância da imagem
                    # img = img.resize(SIZE, PILImage.ANTIALIAS)

                    img.width = IMAGE_WIDTH
                    # img.height = IMAGE_HEIGHT

                    # Definir a posição da imagem no centro da célula (ajuste de deslocamento)
                    cell = ws[f'A{row}']
                    cell_coordinate = cell.coordinate  # Exemplo: "A2"
                    ws.add_image(img, cell_coordinate)  # Inserir imagem na célula

                    # Ajustar altura da linha para exibir corretamente a imagem
                    # ws.row_dimensions[row].height = IMAGE_HEIGHT * 0.85

                    # Remover o texto da célula para manter apenas a imagem
                    cell.value = ""

            seqAnterior = seqAtual

        else:
            if cell_value:
                cell = ws[f'A{row}']
                cell_coordinate = cell.coordinate  # Exemplo: "A2"
                cell.value = ""


    # Criar pasta de saída, se necessário (corrigindo o erro)
    output_dir = os.path.dirname(arquivo_excel)
    if output_dir:  # Só cria se existir um diretório especificado
        os.makedirs(output_dir, exist_ok=True)

    # Salvar o arquivo modificado
    wb.save(arquivo_excel)



def convertCsvToExcel(csvFile, excelFile):
    df = pd.read_csv(csvFile,encoding="utf-8", sep=";")

    df.to_excel(excelFile,index=False, engine="openpyxl")

    ajustarColunas(excelFile)

    adicionarImagensExcel(excelFile)



def main():
    # Testa se o arquivo CSV foi informado
    if len(sys.argv) < 2:
        print("Arquivo CSV não informado.")
        sys.exit(2)

    # Atribui o nome do arquivo CSV a uma variável
    csvFile = sys.argv[1]

    # Verifica se o arquivo CSV existe
    if not os.path.exists(csvFile):
        print(f"Erro: O arquivo '{csvFile}' não foi encontrado.")
        return

    excelFile = csvFile.replace(".csv", ".xlsx")

    convertCsvToExcel(csvFile, excelFile)


if __name__ == "__main__":
    main()  # Executa a função principal
