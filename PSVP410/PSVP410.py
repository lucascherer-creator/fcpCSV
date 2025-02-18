import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from app_location import get_app_location

def adjust_excel_formatting(ws):
    """Adjust column widths and row heights in worksheet."""
    # Adjust column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Adjust row heights
    for row in ws.iter_rows():
        max_height = max(
            (15 * (str(cell.value).count('\n') + 1) if isinstance(cell.value, str) else 15
            for cell in row),
            default=15
        )
        ws.row_dimensions[row[0].row].height = max_height


def add_images_to_excel(ws, image_width=200):
    """ Variável para o local da aplicação """
    paths = get_app_location()
    image_folder = paths['app_dir'] + "..\\IMAGENS\\"

    """Add images to Excel worksheet based on file paths in column A."""
    ws.column_dimensions['A'].width = 30
    prev_seq = None

    for row in range(2, ws.max_row + 1):
        curr_seq = ws[f'D{row}'].value
        if curr_seq != prev_seq:
            cell = ws[f'A{row}']
            img_path = image_folder + (cell.value and cell.value.strip())

            if img_path and os.path.exists(img_path):
                img = Image(img_path)
                img.width = image_width
                ws.add_image(img, cell.coordinate)
                cell.value = ""

            prev_seq = curr_seq
        else:
            ws[f'A{row}'].value = ""

def convert_csv_to_excel(csv_file, excel_file):
    """Convert CSV to Excel and process images."""
    # Convert CSV to Excel
    pd.read_csv(csv_file, encoding="utf-8", sep=";").to_excel(excel_file, index=False, engine="openpyxl")

    # Process Excel file
    wb = load_workbook(excel_file)
    ws = wb.active

    adjust_excel_formatting(ws)
    add_images_to_excel(ws)

    wb.save(excel_file)

def main():
    if len(sys.argv) < 2:
        print("Error: CSV file not specified.")
        sys.exit(2)

    csv_file = sys.argv[1]
    if not os.path.exists(csv_file):
        print(f"Error: File '{csv_file}' not found.")
        sys.exit(1)

    excel_file = csv_file.replace(".csv", ".xlsx")
    convert_csv_to_excel(csv_file, excel_file)

if __name__ == "__main__":
    main()
