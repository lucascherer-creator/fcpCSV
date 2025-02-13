import os
import sys
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def obter_pasta_imagens():
    """
    Retorna o caminho correto da pasta de imagens dependendo de como o script é executado.
    Se for executado como script normal, usa o diretório atual.
    Se for executado como executável criado pelo PyInstaller, usa o diretório extraído temporário.
    """
    if getattr(sys, 'frozen', False):
        # Quando o PyInstaller cria o executável, ele cria uma pasta temporária com os arquivos.
        base_path = sys._MEIPASS  # Diretório temporário onde o PyInstaller extrai os arquivos
    else:
        # Quando executado como script Python normal
        base_path = os.path.abspath(".")  # Diretório atual

    return os.path.join(base_path, "images")  # Diretório "images" dentro do diretório do executável ou script

def criar_arquivo_excel(arquivo_csv, arquivo_excel):
    """
    Função para criar um arquivo Excel a partir de um arquivo CSV.
    """
    with open(arquivo_csv, mode="r") as file:
        reader = csv.reader(file)
        wb = Workbook()  # Cria um novo arquivo Excel
        ws = wb.active  # Acessa a planilha ativa
        
        for row in reader:
            ws.append(row)  # Adiciona cada linha do CSV ao Excel
        
        wb.save(arquivo_excel)  # Salva o arquivo Excel gerado

def adicionar_imagens_excel(arquivo_excel):
    """
    Função para adicionar imagens ao arquivo Excel.
    As imagens são inseridas nas células do Excel conforme os nomes das imagens presentes no arquivo CSV.
    """
    wb = load_workbook(arquivo_excel)  # Abre o arquivo Excel
    ws = wb.active  # Acessa a planilha ativa

    image_folder = obter_pasta_imagens()  # Obtém o caminho da pasta de imagens

    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):  # Itera pelas linhas (começando na segunda linha)
        cell = row[0]  # A primeira célula de cada linha
        nome_imagem = cell.value  # Obtém o valor da célula, que é o nome da imagem

        if isinstance(nome_imagem, str) and nome_imagem.endswith(".bmp"):
            caminho_imagem = os.path.join(image_folder, nome_imagem)  # Cria o caminho completo da imagem
            if os.path.exists(caminho_imagem):  # Verifica se a imagem existe no caminho especificado
                img = Image(caminho_imagem)  # Cria um objeto de imagem
                img.width, img.height = 100, 100  # Ajusta o tamanho da imagem
                ws.add_image(img, cell.coordinate)  # Adiciona a imagem na célula correspondente
                ws.row_dimensions[cell.row].height = 120  # Ajusta a altura da linha para acomodar a imagem

    wb.save(arquivo_excel)  # Salva as alterações no arquivo Excel

def main():
    """
    Função principal que pede ao usuário para fornecer o nome do arquivo CSV,
    cria um arquivo Excel a partir dele e adiciona as imagens conforme necessário.
    """
    # Solicita ao usuário o nome do arquivo CSV
    arquivo_csv = input("Digite o nome do arquivo CSV (com extensão .csv): ")

    # Ajusta o caminho para o arquivo CSV com base na execução do script ou executável
    if getattr(sys, 'frozen', False):
        # Se o script for executado como um executável, ajusta para o diretório temporário
        base_path = sys._MEIPASS  # Diretório temporário onde o PyInstaller extrai os arquivos
        arquivo_csv = os.path.join(base_path, arquivo_csv)
    else:
        # Quando executado como script Python normal
        base_path = os.path.abspath(".")  # Diretório atual
        arquivo_csv = os.path.join(base_path, arquivo_csv)  # Concatena o diretório atual com o nome do arquivo

    # Verifica se o arquivo CSV existe
    if not os.path.exists(arquivo_csv):
        print(f"O arquivo {arquivo_csv} não foi encontrado.")
        return  # Encerra o programa se o arquivo não for encontrado

    # Nome do arquivo Excel de saída será o mesmo nome do CSV, mas com extensão .xlsx
    arquivo_excel = arquivo_csv.replace(".csv", ".xlsx")

    # Criação do arquivo Excel a partir do CSV, se ele não existir
    if not os.path.exists(arquivo_excel):
        print(f"Arquivo Excel {arquivo_excel} não encontrado. Criando um novo arquivo...")
        criar_arquivo_excel(arquivo_csv, arquivo_excel)  # Cria o arquivo Excel a partir do CSV
    
    # Agora, adicionamos as imagens no arquivo Excel
    adicionar_imagens_excel(arquivo_excel)

    print(f"Imagens adicionadas ao arquivo {arquivo_excel}")

if __name__ == "__main__":
    main()  # Executa a função principal
